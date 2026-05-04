from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


def _to_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    return int(value)


def _styles(runtime: dict[str, Any]) -> dict[str, Any]:
    colors = runtime["xlsx"]
    return {
        "title_fill": PatternFill("solid", fgColor=colors["primary_color"]),
        "header_fill": PatternFill("solid", fgColor=colors["secondary_color"]),
        "light_fill": PatternFill("solid", fgColor=colors["light_fill"]),
        "soft_fill": PatternFill("solid", fgColor=colors["soft_fill"]),
        "zebra_fill": PatternFill("solid", fgColor="F8FAFC"),
        "title_font": Font(name="Arial", bold=True, color="FFFFFF", size=14),
        "header_font": Font(name="Arial", bold=True, color="FFFFFF"),
        "bold_font": Font(name="Arial", bold=True, color=colors["text_color"]),
        "body_font": Font(name="Arial", color=colors["text_color"]),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


def _auto_width(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _title(ws, cell_range: str, text: str, styles: dict[str, Any]) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = styles["title_fill"]
    cell.font = styles["title_font"]
    cell.alignment = styles["center"]


def _header_row(ws, row_idx: int, values: list[str], styles: dict[str, Any]) -> None:
    for column_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=column_idx, value=value)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]


def _body_row(ws, row_idx: int, values: list[Any], styles: dict[str, Any], left_columns: set[int] | None = None) -> None:
    left_columns = left_columns or set()
    for column_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=column_idx, value=value)
        cell.font = styles["body_font"]
        cell.alignment = styles["left"] if column_idx in left_columns else styles["center"]
        if row_idx % 2 == 0:
            cell.fill = styles["zebra_fill"]


def _sum_hours(rows: list[dict[str, Any]], key: str) -> int:
    return sum(_to_int(row.get(key)) for row in rows)


def build_backlog(scope_context: dict[str, Any], answers: dict[str, Any], output_path: Path, runtime: dict[str, Any]) -> Path:
    styles = _styles(runtime)
    workbook = Workbook()

    sprint_name = answers["sprint_name"]
    sprint_period = answers["sprint_period"]
    sprint_goal = answers["sprint_goal"]
    hourly_rate = _to_int(answers.get("hourly_rate"))
    backlog_items = answers["backlog_items"]
    allocation_rows = answers["allocation_rows"]
    absences = answers["absences"]
    risks = answers.get("risk_items", [])

    total_hours = _sum_hours(backlog_items, "hours")
    total_cost = total_hours * hourly_rate
    total_absence_hours = _sum_hours(absences, "hours")

    summary = workbook.active
    summary.title = "Resumo"
    _title(summary, "A1:J1", f"Backlog da {sprint_name}", styles)
    summary["A3"] = "Projeto"
    summary["B3"] = answers["project_name"]
    summary["A4"] = "Período"
    summary["B4"] = sprint_period
    summary["A5"] = "Objetivo"
    summary["B5"] = sprint_goal
    summary["E3"] = "Tarefas"
    summary["F3"] = len(backlog_items)
    summary["E4"] = "Horas"
    summary["F4"] = total_hours
    summary["E5"] = "Custo"
    summary["F5"] = total_cost
    summary["E6"] = "Ausências"
    summary["F6"] = total_absence_hours
    for cell_ref in ("A3", "A4", "A5", "E3", "E4", "E5", "E6"):
        summary[cell_ref].font = styles["bold_font"]
    for cell_ref in ("B3", "B4", "B5", "F3", "F4", "F5", "F6"):
        summary[cell_ref].font = styles["body_font"]
    summary["F5"].number_format = 'R$ #,##0'
    summary["A8"] = "Riscos da sprint"
    summary["A8"].font = styles["bold_font"]
    risk_start = 9
    if risks:
        for idx, risk in enumerate(risks, start=risk_start):
            summary[f"A{idx}"] = f"- {risk}"
            summary[f"A{idx}"].alignment = styles["left"]
    else:
        summary["A9"] = "- Nenhum risco informado."
    _auto_width(summary, {"A": 18, "B": 50, "E": 14, "F": 14})

    allocation_sheet = workbook.create_sheet("Alocacao")
    _title(allocation_sheet, "A1:G1", f"Equipe Alocada - {sprint_name}", styles)
    _header_row(
        allocation_sheet,
        3,
        ["Nome", "Papel", "Capacidade planejada (h)", "Horas alocadas", "Folga (h)", "Utilização (%)", "Observação"],
        styles,
    )
    for row_idx, row in enumerate(allocation_rows, start=4):
        capacity = _to_int(row.get("capacity_hours"))
        allocated = _to_int(row.get("allocated_hours"))
        slack = capacity - allocated
        utilization = round((allocated / capacity) * 100, 1) if capacity else 0
        _body_row(
            allocation_sheet,
            row_idx,
            [row.get("name", ""), row.get("role", ""), capacity, allocated, slack, utilization, row.get("availability_note", "")],
            styles,
            left_columns={1, 2, 7},
        )
    total_row = len(allocation_rows) + 4
    allocation_sheet.cell(row=total_row, column=2, value="Totais").font = styles["bold_font"]
    allocation_sheet.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row - 1})").font = styles["bold_font"]
    allocation_sheet.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row - 1})").font = styles["bold_font"]
    allocation_sheet.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row - 1})").font = styles["bold_font"]
    allocation_sheet.cell(row=total_row, column=6, value=f"=IF(C{total_row}=0,0,(D{total_row}/C{total_row})*100)").font = styles["bold_font"]
    allocation_sheet.cell(row=total_row, column=6).number_format = "0.0"
    _auto_width(allocation_sheet, {"A": 26, "B": 26, "C": 20, "D": 16, "E": 12, "F": 14, "G": 34})

    backlog = workbook.create_sheet("Backlog")
    _title(backlog, "A1:K1", f"Itens do Backlog - {sprint_name}", styles)
    _header_row(
        backlog,
        3,
        ["ID", "Épico", "Tarefa", "Responsável", "Início", "Fim", "Horas", "Status", "Impacto", "Mitigação", "Critério de aceite"],
        styles,
    )
    for row_idx, row in enumerate(backlog_items, start=4):
        values = [
            row.get("id", ""),
            row.get("epic", ""),
            row.get("task", ""),
            row.get("owner", ""),
            row.get("start_date", ""),
            row.get("end_date", ""),
            _to_int(row.get("hours")),
            row.get("status", ""),
            row.get("impact", ""),
            row.get("mitigation", ""),
            row.get("acceptance", ""),
        ]
        _body_row(backlog, row_idx, values, styles, left_columns={2, 3, 4, 9, 10, 11})
    total_backlog_row = len(backlog_items) + 4
    backlog.cell(row=total_backlog_row, column=6, value="Total").font = styles["bold_font"]
    backlog.cell(row=total_backlog_row, column=7, value=f"=SUM(G4:G{total_backlog_row - 1})").font = styles["bold_font"]
    _auto_width(backlog, {"A": 12, "B": 18, "C": 36, "D": 22, "E": 14, "F": 14, "G": 10, "H": 18, "I": 26, "J": 26, "K": 28})

    absences_sheet = workbook.create_sheet("Ausencias")
    _title(absences_sheet, "A1:D1", f"Ausências - {sprint_name}", styles)
    _header_row(absences_sheet, 3, ["Nome", "Data", "Horas", "Impacto"], styles)
    for row_idx, row in enumerate(absences, start=4):
        _body_row(
            absences_sheet,
            row_idx,
            [row.get("name", ""), row.get("date", ""), _to_int(row.get("hours")), row.get("impact", "")],
            styles,
            left_columns={1, 4},
        )
    _auto_width(absences_sheet, {"A": 24, "B": 14, "C": 10, "D": 48})

    risks_sheet = workbook.create_sheet("Riscos")
    _title(risks_sheet, "A1:B1", f"Riscos da Sprint - {sprint_name}", styles)
    _header_row(risks_sheet, 3, ["ID", "Descrição"], styles)
    for row_idx, risk in enumerate(risks, start=4):
        _body_row(risks_sheet, row_idx, [f"R{row_idx - 3:02d}", risk], styles, left_columns={2})
    _auto_width(risks_sheet, {"A": 10, "B": 88})

    chart = BarChart()
    chart.title = "Horas alocadas por colaborador"
    chart.y_axis.title = "Horas"
    chart.height = 7
    chart.width = 15
    data = Reference(allocation_sheet, min_col=4, max_col=4, min_row=3, max_row=max(4, len(allocation_rows) + 3))
    categories = Reference(allocation_sheet, min_col=1, min_row=4, max_row=max(4, len(allocation_rows) + 3))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    summary.add_chart(chart, "E8")

    summary.freeze_panes = "A3"
    allocation_sheet.freeze_panes = "A3"
    backlog.freeze_panes = "A3"
    absences_sheet.freeze_panes = "A3"
    risks_sheet.freeze_panes = "A3"

    workbook.save(output_path)
    return output_path


def build_cycle_details(scope_context: dict[str, Any], answers: dict[str, Any], output_path: Path, runtime: dict[str, Any]) -> Path:
    styles = _styles(runtime)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Sprints"
    _title(ws, "A1:I1", "Detalhamento dos Ciclos", styles)
    _header_row(
        ws,
        3,
        ["Sprint", "Período", "Objetivo", "Horas Estimadas", "Horas Gastas", "Diferença", "Acertabilidade (%)", "Custo por Sprint (R$)", "Observações"],
        styles,
    )
    cycles = answers["cycles"]
    hourly_rate = _to_int(answers["hourly_rate"])
    for row_idx, row in enumerate(cycles, start=4):
        estimated = _to_int(row.get("estimated_hours"))
        spent = _to_int(row.get("spent_hours"))
        values = [
            row.get("name", ""),
            row.get("period", ""),
            row.get("objective", ""),
            estimated,
            spent,
            spent - estimated,
            round((estimated / spent) * 100, 1) if spent else 0,
            spent * hourly_rate,
            row.get("notes", ""),
        ]
        _body_row(ws, row_idx, values, styles, left_columns={2, 3, 9})
        ws.cell(row=row_idx, column=8).number_format = 'R$ #,##0'
    total_row = len(cycles) + 4
    ws.cell(row=total_row, column=3, value="Totais").font = styles["bold_font"]
    ws.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row - 1})").font = styles["bold_font"]
    ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row - 1})").font = styles["bold_font"]
    ws.cell(row=total_row, column=6, value=f"=SUM(F4:F{total_row - 1})").font = styles["bold_font"]
    ws.cell(row=total_row, column=7, value=f"=IF(E{total_row}=0,0,(D{total_row}/E{total_row})*100)").font = styles["bold_font"]
    ws.cell(row=total_row, column=8, value=f"=SUM(H4:H{total_row - 1})").font = styles["bold_font"]
    ws.cell(row=total_row, column=8).number_format = 'R$ #,##0'
    _auto_width(ws, {"A": 14, "B": 24, "C": 34, "D": 16, "E": 14, "F": 12, "G": 18, "H": 19, "I": 42})

    chart_sheet = workbook.create_sheet("Gráficos")
    _title(chart_sheet, "A1:H1", "Indicadores por Sprint", styles)

    bar = BarChart()
    bar.title = "Horas estimadas vs horas gastas"
    data = Reference(ws, min_col=4, max_col=5, min_row=3, max_row=len(cycles) + 3)
    categories = Reference(ws, min_col=1, min_row=4, max_row=len(cycles) + 3)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(categories)
    bar.height = 8
    bar.width = 18
    chart_sheet.add_chart(bar, "A3")

    bar2 = BarChart()
    bar2.title = "Acertabilidade (%)"
    data2 = Reference(ws, min_col=7, max_col=7, min_row=3, max_row=len(cycles) + 3)
    bar2.add_data(data2, titles_from_data=True)
    bar2.set_categories(categories)
    bar2.height = 8
    bar2.width = 18
    chart_sheet.add_chart(bar2, "A20")

    ws.freeze_panes = "A3"
    workbook.save(output_path)
    return output_path
